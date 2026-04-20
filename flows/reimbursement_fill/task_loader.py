from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from flows.reimbursement_fill.task_model import (
    INVOICE_SHEET_REQUIRED_FIELDS,
    TASK_SHEET_REQUIRED_FIELDS,
    ReimbursementInvoiceRecord,
    ReimbursementTaskRecord,
)


TASK_DESCRIPTION_ROW = {
    "task_id": "任务唯一编号",
    "business_department": "业务单位",
}

INVOICE_DESCRIPTION_ROW = {
    "task_id": "任务唯一编号",
    "company_count": "公司人数",
    "approved_amount": "核定报账金额",
}


def load_tasks(task_path: str) -> list[ReimbursementTaskRecord]:
    path = Path(task_path).resolve()
    if path.suffix.lower() != ".xlsx":
        raise RuntimeError("reimbursement-fill 当前仅支持 .xlsx 任务文件")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        task_ws = wb["task"]
        invoice_ws = wb["invoice"]
    except KeyError as exc:
        raise RuntimeError("任务文件必须包含 `task` 与 `invoice` 两个 sheet") from exc

    task_rows = list(task_ws.iter_rows(values_only=True))
    invoice_rows = list(invoice_ws.iter_rows(values_only=True))

    task_headers = _normalize_headers(task_rows[0] if task_rows else None)
    invoice_headers = _normalize_headers(invoice_rows[0] if invoice_rows else None)

    tasks_by_id: dict[str, ReimbursementTaskRecord] = {}
    ordered_tasks: list[ReimbursementTaskRecord] = []

    for row_index, row in enumerate(task_rows[1:], start=2):
        row_dict = _row_to_dict(task_headers, row)
        if _is_task_description_row(row_dict):
            continue
        task_id = row_dict.get("task_id", "")
        if not task_id:
            continue
        task = ReimbursementTaskRecord(
            task_id=task_id,
            business_department=row_dict.get("business_department", ""),
            payment_purpose=row_dict.get("payment_purpose", ""),
            source_row=row_index,
        )
        tasks_by_id[task.task_id] = task
        ordered_tasks.append(task)

    for row_index, row in enumerate(invoice_rows[1:], start=2):
        row_dict = _row_to_dict(invoice_headers, row)
        if _is_invoice_description_row(row_dict):
            continue
        task_id = row_dict.get("task_id", "")
        if not task_id:
            continue
        invoice = ReimbursementInvoiceRecord(
            task_id=task_id,
            company_count=row_dict.get("company_count", ""),
            file_path=row_dict.get("file_path", ""),
            remark=row_dict.get("remark", ""),
            source_row=row_index,
            approved_amount=row_dict.get("approved_amount", ""),
        )
        if task_id in tasks_by_id:
            tasks_by_id[task_id].invoices.append(invoice)

    return ordered_tasks


def validate_tasks(tasks: list[ReimbursementTaskRecord], task_path: str | None = None) -> list[str]:
    errors: list[str] = []
    seen_ids: set[str] = set()

    if task_path:
        errors.extend(_validate_workbook_structure(Path(task_path).resolve()))

    for task in tasks:
        for field in TASK_SHEET_REQUIRED_FIELDS:
            value = getattr(task, field)
            if not value:
                errors.append(f"row {task.source_row} [task]: `{field}` is required")
        if task.task_id in seen_ids:
            errors.append(f"row {task.source_row} [task]: duplicated `task_id` `{task.task_id}`")
        seen_ids.add(task.task_id)
        if not task.invoices:
            errors.append(f"row {task.source_row} [task]: no invoice rows found for task_id `{task.task_id}`")
        for invoice in task.invoices:
            for field in INVOICE_SHEET_REQUIRED_FIELDS:
                value = getattr(invoice, field)
                if not value:
                    errors.append(f"row {invoice.source_row} [invoice]: `{field}` is required")
            if invoice.company_count:
                try:
                    int(invoice.company_count)
                except ValueError:
                    errors.append(f"row {invoice.source_row} [invoice]: `company_count` must be an integer")
            if invoice.approved_amount:
                try:
                    float(invoice.approved_amount)
                except ValueError:
                    errors.append(f"row {invoice.source_row} [invoice]: `approved_amount` must be a number")
    return errors


def _validate_workbook_structure(path: Path) -> list[str]:
    errors: list[str] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        task_ws = wb["task"]
        invoice_ws = wb["invoice"]
    except KeyError as exc:
        return ["任务文件必须包含 `task` 与 `invoice` 两个 sheet"]

    task_headers = _normalize_headers(next(task_ws.iter_rows(min_row=1, max_row=1, values_only=True), None))
    invoice_headers = _normalize_headers(next(invoice_ws.iter_rows(min_row=1, max_row=1, values_only=True), None))

    for field in (*TASK_SHEET_REQUIRED_FIELDS, "upload_result"):
        if field not in task_headers:
            errors.append(f"sheet `task`: missing column `{field}`")
    for field in INVOICE_SHEET_REQUIRED_FIELDS:
        if field not in invoice_headers:
            errors.append(f"sheet `invoice`: missing column `{field}`")
    return errors


def _normalize_headers(row: tuple | list | None) -> list[str]:
    if not row:
        return []
    return [str(value).strip() if value is not None else "" for value in row]


def _row_to_dict(headers: list[str], row: tuple | list | None) -> dict[str, str]:
    values = list(row or [])
    result: dict[str, str] = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        value = values[idx] if idx < len(values) else ""
        result[header] = str(value).strip() if value is not None else ""
    return result


def _is_task_description_row(row: dict[str, str]) -> bool:
    return (
        row.get("task_id", "") == TASK_DESCRIPTION_ROW["task_id"]
        and row.get("business_department", "") == TASK_DESCRIPTION_ROW["business_department"]
    )


def _is_invoice_description_row(row: dict[str, str]) -> bool:
    return (
        row.get("task_id", "") == INVOICE_DESCRIPTION_ROW["task_id"]
        and row.get("company_count", "") == INVOICE_DESCRIPTION_ROW["company_count"]
    )


__all__ = ["load_tasks", "validate_tasks"]
