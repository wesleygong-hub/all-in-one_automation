from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


class TaskResultWriteError(RuntimeError):
    pass


def write_task_result(task_path: str, task, result_text: str) -> str:
    path = Path(task_path).resolve()
    try:
        if path.suffix.lower() == ".xlsx":
            _write_xlsx_result(path, task, result_text)
        else:
            rows = _load_task_rows(path)
            _apply_result_to_csv_rows(rows, task, result_text)
            _write_task_rows(path, rows)
        return str(path)
    except PermissionError as exc:
        fallback_path = path.with_name(f"{path.stem}.result{path.suffix}")
        try:
            if path.suffix.lower() == ".xlsx":
                _write_xlsx_result(path, task, result_text, target_path=fallback_path)
            else:
                rows = _load_task_rows(path)
                _apply_result_to_csv_rows(rows, task, result_text)
                _write_task_rows(fallback_path, rows)
            raise TaskResultWriteError(
                f"任务结果无法写回原文件，可能正被占用；已写入备用文件: {fallback_path}"
            ) from exc
        except Exception as fallback_exc:
            raise TaskResultWriteError(
                f"任务结果无法写回原文件，且备用文件写入失败: {fallback_path}"
            ) from fallback_exc


def _load_task_rows(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.reader(fh))


def _apply_result_to_csv_rows(rows: list[list[str]], task, result_text: str) -> None:
    target_index = getattr(task, "source_row", 0) - 1
    if target_index < 0 or target_index >= len(rows):
        raise RuntimeError(f"task row out of range: {getattr(task, 'source_row', None)}")

    row = rows[target_index]
    while len(row) < 7:
        row.append("")
    row[6] = result_text

    if rows:
        while len(rows[0]) < 7:
            rows[0].append("")
        rows[0][6] = "upload_result"
    if len(rows) > 1:
        while len(rows[1]) < 7:
            rows[1].append("")
        if not rows[1][6]:
            rows[1][6] = "上传结果"


def _write_task_rows(path: Path, rows: list[list[str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerows(rows)


def _write_xlsx_result(path: Path, task, result_text: str, target_path: Path | None = None) -> None:
    wb = load_workbook(path)
    ws = wb["task"] if "task" in wb.sheetnames else wb.active

    header_row = [cell.value for cell in ws[1]]
    result_col = _find_or_create_result_column(ws, header_row)
    target_row = getattr(task, "source_row", None)
    if not target_row:
        raise RuntimeError("task source_row is required for xlsx write-back")
    ws.cell(row=target_row, column=result_col, value=result_text)

    if ws.max_row >= 2 and not ws.cell(row=2, column=result_col).value:
        ws.cell(row=2, column=result_col, value="结果回写")

    wb.save(target_path or path)


def _find_or_create_result_column(ws, header_row: list[object | None]) -> int:
    for idx, value in enumerate(header_row, start=1):
        if str(value).strip() == "upload_result":
            return idx
    new_col = len(header_row) + 1 if header_row else 1
    ws.cell(row=1, column=new_col, value="upload_result")
    return new_col


__all__ = ["TaskResultWriteError", "write_task_result"]
